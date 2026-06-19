"""Parse AKS bank settlement files (XLS preferred, PDF fallback)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field

AKS_COURIER_FEE_RSD = 490.0
ORDER_ID_RE = re.compile(r"917\d{11}")
SETTLEMENT_REF_RE = re.compile(r"Specifikacija[:\s]+(\d+)", re.I)
SETTLEMENT_DATE_RE = re.compile(r"na dan\s+(\d{2}\.\d{2}\.\d{4})", re.I)
# Serbian PDFs: 1650.00, 1650,00, or 1.650,00
PDF_AMOUNT_RE = r"(?:\d{1,3}(?:\.\d{3})*,\d{2}|\d+[.,]\d{2})"
PDF_LINE_RE = re.compile(
    rf"(917\d{{11}}).*?({PDF_AMOUNT_RE})",
)


@dataclass
class SettlementLine:
    order_id: str
    payer_name: str
    aks_amount_rsd: float
    confirmed_at: str | None = None


@dataclass
class ParsedSettlement:
    settlement_ref: str
    settlement_date: str | None
    filename: str
    lines: list[SettlementLine] = field(default_factory=list)

    @property
    def total_aks_rsd(self) -> float:
        return round(sum(line.aks_amount_rsd for line in self.lines), 2)


def settlement_product_rsd(
    aks_amount_rsd: float,
    *,
    bundle_count: int = 0,
    sale_unit_rsd: float = 1000.0,
) -> float:
    """Credit product revenue only — ignore AKS courier fee (490 RSD)."""
    unit = max(1.0, float(sale_unit_rsd or 1000.0))
    if bundle_count > 0:
        return round(bundle_count * unit, 2)
    inferred_bundles = max(1, round((float(aks_amount_rsd) - AKS_COURIER_FEE_RSD) / unit))
    return round(inferred_bundles * unit, 2)


def _normalize_order_id(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) == 14 and digits.startswith("917"):
        return digits
    return ""


def _ref_from_filename(filename: str) -> str | None:
    match = re.search(r"S(\d+)", filename or "", re.I)
    return match.group(1) if match else None


def _detect_file_ext(filename: str, content: bytes) -> str:
    name = (filename or "").lower()
    if content[:4] == b"%PDF":
        return "pdf"
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    if content[:2] == b"PK":
        return "xlsx"
    if "." in name:
        return name.rsplit(".", 1)[-1]
    return ""


def _parse_amount(raw: str) -> float | None:
    text = (raw or "").strip()
    if not text:
        return None
    if "," in text and "." in text:
        normalized = text.replace(".", "").replace(",", ".")
    elif "," in text:
        normalized = text.replace(",", ".")
    else:
        normalized = text
    try:
        return float(normalized)
    except ValueError:
        return None


def parse_settlement_file(filename: str, content: bytes) -> ParsedSettlement:
    ext = _detect_file_ext(filename, content)
    if ext == "xls":
        parsed = _parse_xls(content, filename)
    elif ext == "xlsx":
        parsed = _parse_xlsx(content, filename)
    elif ext == "pdf":
        parsed = _parse_pdf(content, filename)
    else:
        raise ValueError("Unsupported file type. Upload the AKS .xls, .xlsx, or .pdf settlement.")

    if not parsed.lines:
        raise ValueError("No Order IDs found in this file.")
    if not parsed.settlement_ref:
        parsed.settlement_ref = _ref_from_filename(filename) or "unknown"
    return parsed


def _parse_xls(content: bytes, filename: str) -> ParsedSettlement:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = [
        [str(sheet.cell_value(row_idx, col)).strip() for col in range(sheet.ncols)]
        for row_idx in range(sheet.nrows)
    ]
    return _parse_spreadsheet_rows(rows, filename)


def _parse_xlsx(content: bytes, filename: str) -> ParsedSettlement:
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows = [
        [str(v).strip() if v is not None else "" for v in row]
        for row in book.active.iter_rows(values_only=True)
    ]
    return _parse_spreadsheet_rows(rows, filename)


def _parse_spreadsheet_rows(rows: list[list[str]], filename: str) -> ParsedSettlement:
    settlement_ref = _ref_from_filename(filename) or ""
    settlement_date: str | None = None
    lines: list[SettlementLine] = []
    header_row = -1

    for row_idx, cells in enumerate(rows):
        joined = " ".join(cells)
        if not settlement_ref:
            ref_match = SETTLEMENT_REF_RE.search(joined)
            if ref_match:
                settlement_ref = ref_match.group(1)
        if not settlement_date:
            date_match = SETTLEMENT_DATE_RE.search(joined)
            if date_match:
                settlement_date = date_match.group(1)
        if "NalogID" in cells and "Iznos" in cells:
            header_row = row_idx
            break

    if header_row < 0:
        raise ValueError("Could not read AKS spreadsheet header (NalogID / Iznos).")

    headers = rows[header_row]
    col = {name: idx for idx, name in enumerate(headers) if name}

    for cells in rows[header_row + 1:]:
        if len(cells) <= max(col.get("NalogID", 1), col.get("Iznos", 4)):
            continue
        order_id = _normalize_order_id(cells[col.get("NalogID", 1)])
        if not order_id:
            continue
        try:
            aks_amount = float(cells[col.get("Iznos", 4)])
        except (TypeError, ValueError):
            continue
        if aks_amount <= 0:
            continue
        payer = cells[col.get("Platilac", 5)] if col.get("Platilac") is not None else ""
        confirmed = cells[col.get("Vreme potvrde", 0)] if col.get("Vreme potvrde") is not None else ""
        lines.append(
            SettlementLine(
                order_id=order_id,
                payer_name=payer.strip(),
                aks_amount_rsd=round(aks_amount, 2),
                confirmed_at=confirmed.strip() or None,
            )
        )

    return ParsedSettlement(
        settlement_ref=settlement_ref,
        settlement_date=settlement_date,
        filename=filename,
        lines=lines,
    )


def _pdf_line_from_match(raw_line: str, order_id: str, amount_raw: str) -> SettlementLine | None:
    aks_amount = _parse_amount(amount_raw)
    if aks_amount is None or aks_amount <= 0:
        return None
    payer = ""
    after_amount = raw_line.split(amount_raw, 1)
    if len(after_amount) > 1:
        payer = after_amount[1].strip().split("  ")[0].strip()
    return SettlementLine(
        order_id=order_id,
        payer_name=payer,
        aks_amount_rsd=round(aks_amount, 2),
    )


def _extract_pdf_lines(text: str) -> list[SettlementLine]:
    lines: list[SettlementLine] = []
    seen: set[str] = set()

    for raw_line in text.splitlines():
        match = PDF_LINE_RE.search(raw_line)
        if not match:
            continue
        order_id = match.group(1)
        if order_id in seen:
            continue
        row = _pdf_line_from_match(raw_line, order_id, match.group(2))
        if not row:
            continue
        seen.add(order_id)
        lines.append(row)

    if lines:
        return lines

    # Some AKS PDFs break rows across lines — pair each ID with the next amount in the text.
    amount_re = re.compile(PDF_AMOUNT_RE)
    for order_id in ORDER_ID_RE.findall(text):
        if order_id in seen:
            continue
        start = text.find(order_id)
        if start < 0:
            continue
        window = text[start : start + 120]
        amount_match = amount_re.search(window[len(order_id) :])
        if not amount_match:
            continue
        aks_amount = _parse_amount(amount_match.group(0))
        if aks_amount is None or aks_amount <= 0:
            continue
        seen.add(order_id)
        lines.append(
            SettlementLine(
                order_id=order_id,
                payer_name="",
                aks_amount_rsd=round(aks_amount, 2),
            )
        )
    return lines


def _parse_pdf(content: bytes, filename: str) -> ParsedSettlement:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError(
            "PDF support is missing on the server. Upload the .xls or .xlsx settlement instead."
        ) from exc

    reader = PdfReader(io.BytesIO(content), strict=False)
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    if not text.strip():
        raise ValueError(
            "This PDF has no readable text (scanned image). "
            "Download the .xls or .xlsx specifikacija from AKS email instead."
        )

    settlement_ref = ""
    ref_match = SETTLEMENT_REF_RE.search(text)
    if ref_match:
        settlement_ref = ref_match.group(1)
    if not settlement_ref:
        settlement_ref = _ref_from_filename(filename) or ""

    date_match = SETTLEMENT_DATE_RE.search(text)
    settlement_date = date_match.group(1) if date_match else None
    lines = _extract_pdf_lines(text)

    return ParsedSettlement(
        settlement_ref=settlement_ref,
        settlement_date=settlement_date,
        filename=filename,
        lines=lines,
    )
